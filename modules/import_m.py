from openpyxl import Workbook, load_workbook
import csv

def importExcel(fileDialog, window, title):
    filepath, _ = fileDialog.getOpenFileName(window, 'Open File', './', "Table (*.xlsx)")
    if filepath != "":
        wbSearch = Workbook()
        wbSearch = load_workbook(filepath, data_only=True)
        wsSearch = wbSearch.active
        for i in range(1,150):
            value=wsSearch.cell(row=i, column=2).value
            if value and title in value:
                values = [wsSearch.cell(row=i, column=j).value for j in range(1,55)]
                return values

def importCsv(fileDialog, window):
    filepath, _ = fileDialog.getOpenFileName(window, 'Open File', './', "Table (*.csv)")
    if filepath != "":
        values = []
        with open(filepath, 'r', encoding="utf8") as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=';')
            for row in csv_reader:
                values.append(row)
        return values